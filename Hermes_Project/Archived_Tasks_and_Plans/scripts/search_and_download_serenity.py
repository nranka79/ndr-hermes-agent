import os
import sys
import io
import openpyxl
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# Force UTF-8 output
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

TOKEN_PATH = r"C:\Users\ruhaan\AntiGravity\workspace_power_token.json"

def get_drive_service():
    if not os.path.exists(TOKEN_PATH):
        print("ERROR: Power Token not found.")
        sys.exit(1)
    creds = Credentials.from_authorized_user_file(TOKEN_PATH)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return build('drive', 'v3', credentials=creds)

def download_file(service, file_id, file_name):
    file_metadata = service.files().get(fileId=file_id).execute()
    mime_type = file_metadata.get('mimeType')
    
    if mime_type == 'application/vnd.google-apps.spreadsheet':
        request = service.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if not file_name.endswith('.xlsx'):
            file_name += '.xlsx'
    else:
        request = service.files().get_media(fileId=file_id)

    fh = io.FileIO(file_name, 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
    print(f"Downloaded {file_name}")
    return file_name

def main():
    service = get_drive_service()
    
    query = "name contains 'Serenity Hillview' and name contains 'inventory'"
    results = service.files().list(q=query, fields="files(id, name, mimeType)").execute()
    files = results.get('files', [])
    
    if not files:
        print("No file found.")
        return

    target = files[0]
    print(f"Found file: {target['name']} ({target['id']})")
    
    local_file = download_file(service, target['id'], "serenity_data.xlsx")
    
    wb = openpyxl.load_workbook(local_file, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    # User said "analysis" sheet
    sheet_name = [s for s in wb.sheetnames if 'analysis' in s.lower()]
    if not sheet_name:
        print("Could not find analysis sheet. Using active sheet.")
        ws = wb.active
    else:
        ws = wb[sheet_name[0]]
    
    print(f"\nAnalyzing sheet: {ws.title}")
    
    with open("serenity_inspect.txt", "w", encoding="utf-8") as f:
        for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
            f.write(str(row) + "\n")
    print("Saved first 100 rows to serenity_inspect.txt")

if __name__ == '__main__':
    main()
